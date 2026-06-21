"""
Ведомость оборудования → XLSX (openpyxl).

Колонки: № п/п, Наименование, Марка/Тип, Кол-во, Ед., Примечание.
Позиции извлекаются из текста проекта моделью (structured output) с откатом
на простую эвристику по строкам — чтобы файл формировался и без LLM.
"""
import io
import logging
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from .common import ProjectExportData, ORG_NAME

log = logging.getLogger(__name__)

_HEADERS = ["№ п/п", "Наименование", "Марка/Тип", "Кол-во", "Ед.", "Примечание"]
_WIDTHS = [8, 42, 20, 10, 10, 30]


class EquipmentRow(BaseModel):
    """Одна позиция ведомости оборудования."""
    name: str
    mark: str = ""
    qty: str = ""
    unit: str = "шт"
    note: str = ""


class EquipmentList(BaseModel):
    """Список позиций оборудования (для structured output)."""
    items: list[EquipmentRow] = []


# Счётные единицы для эвристики (без линейных/площадных — это размеры, не оборудование).
_UNIT_RE = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*(шт\.?|ед\.?|компл\.?|комплект\w*|кВт)",
    re.IGNORECASE,
)


async def _extract_via_llm(data: ProjectExportData) -> list[EquipmentRow]:
    """Извлечь позиции оборудования моделью. Возвращает [] при любой ошибке."""
    if not data.raw_text.strip():
        return []
    try:
        from graph.graph import _llm  # единая точка доступа к модели (правило 1)

        system = (
            "Ты — сметчик. Из текста проектных решений извлеки перечень оборудования "
            "в виде структурированного списка позиций. Для каждой позиции: наименование, "
            "марка/тип (если есть), количество, единица измерения, примечание. "
            "Не выдумывай позиции, которых нет в тексте."
        )
        from langchain_core.messages import SystemMessage, HumanMessage

        llm = _llm(streaming=False).with_structured_output(EquipmentList)
        result: EquipmentList = await llm.ainvoke([
            SystemMessage(content=system),
            HumanMessage(content=data.raw_text[:8000]),
        ])
        return result.items or []
    except Exception:
        log.info("Извлечение оборудования через LLM недоступно — откат на эвристику", exc_info=True)
        return []


def _extract_heuristic(data: ProjectExportData) -> list[EquipmentRow]:
    """Простой откат: строки с количеством и единицей измерения → позиции."""
    rows: list[EquipmentRow] = []
    seen: set[str] = set()
    for line in data.raw_text.splitlines():
        line = line.strip(" •-—\t")
        if not line or len(line) > 200:
            continue
        m = _UNIT_RE.search(line)
        if not m:
            continue
        name = line[: m.start()].strip(" :—-") or line
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        rows.append(EquipmentRow(
            name=name[:120],
            qty=m.group(1).replace(",", "."),
            unit=m.group(2).rstrip("."),
        ))
        if len(rows) >= 100:
            break
    return rows


async def build_equipment_sheet(data: ProjectExportData) -> bytes:
    """Сформировать ведомость оборудования и вернуть XLSX как bytes."""
    # Приоритет — детерминированно подобранное оборудование (из расчётного ядра).
    if data.equipment:
        items = [EquipmentRow(name=e["name"], mark=e.get("role", ""),
                              qty=str(e.get("qty", "")), unit="шт",
                              note=e.get("capacity", "")) for e in data.equipment]
    else:
        items = await _extract_via_llm(data)
        if not items:
            items = _extract_heuristic(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость оборудования"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Шапка документа
    ws.merge_cells("A1:F1")
    ws["A1"] = "ВЕДОМОСТЬ ОБОРУДОВАНИЯ"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.merge_cells("A2:F2")
    ws["A2"] = f"{ORG_NAME} · проект «{data.name}» · {data.date_str}"
    ws["A2"].alignment = center
    ws["A2"].font = Font(italic=True, size=10)

    # Заголовки таблицы
    header_row = 4
    for col, title in enumerate(_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Данные
    r = header_row + 1
    if items:
        for i, item in enumerate(items, 1):
            values = [i, item.name, item.mark, item.qty, item.unit, item.note]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = border
                cell.alignment = center if col in (1, 4, 5) else left
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        note = ws.cell(row=r, column=1, value="Нет данных об оборудовании в проекте.")
        note.alignment = center
        note.font = Font(italic=True, color="888888")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = border

    # Ширины колонок
    for col, width in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
